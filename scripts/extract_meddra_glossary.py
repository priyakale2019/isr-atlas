#!/usr/bin/env python3
"""Regenerate data/glossary.js and Glossary sheet from MedDRA search term workbook."""

from __future__ import annotations

import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(
    os.environ.get(
        "MEDDRA_GLOSSARY_XLSX",
        "/Users/pkale/Downloads/MedDRA search term list - PTs and LLTs.xlsx",
    )
)
OUT_JS = ROOT / "data" / "glossary.js"

# Exact terms to drop from glossary (case-insensitive)
EXCLUDED_TERMS = frozenset(
    t.casefold()
    for t in (
        "inflammation at site of injection",
        "reactions",
        "reaction (nos)",
        "swelling of",
    )
)

# Longest-first British → American (word-boundary replacements)
BRITISH_TO_AMERICAN: list[tuple[str, str]] = [
    ("discolouration", "discoloration"),
    ("discolour", "discolor"),
    ("colouration", "coloration"),
    ("hyperaesthesia", "hyperesthesia"),
    ("hypoaesthesia", "hypoesthesia"),
    ("hyperaesthesia", "hyperesthesia"),
    ("dysaesthesia", "dysesthesia"),
    ("paraesthesia", "paresthesia"),
    ("hypaesthesia", "hypesthesia"),
    ("anaesthesia", "anesthesia"),
    ("haemorrhage", "hemorrhage"),
    ("haematoma", "hematoma"),
    ("ischaemia", "ischemia"),
    ("ischaemic", "ischemic"),
    ("oedema", "edema"),
    ("hyperaemia", "hyperemia"),
    ("hypoaemia", "hypoemia"),
    ("anaemia", "anemia"),
    ("leukaemia", "leukemia"),
    ("oesophagitis", "esophagitis"),
    ("oestrogen", "estrogen"),
    ("tumour", "tumor"),
    ("colour", "color"),
    ("favour", "favor"),
    ("behaviour", "behavior"),
    ("labour", "labor"),
    ("centre", "center"),
    ("fibre", "fiber"),
    ("litre", "liter"),
    ("metre", "meter"),
    ("sulphur", "sulfur"),
    ("paediatric", "pediatric"),
    ("glycaemia", "glycemia"),
    ("septicaemia", "septicemia"),
    ("bacteraemia", "bacteremia"),
    ("uraemia", "uremia"),
    ("uraemic", "uremic"),
    ("oedematous", "edematous"),
]

_BRITISH_PATTERNS = [
    (re.compile(rf"\b{re.escape(brit)}\b", re.IGNORECASE), amer)
    for brit, amer in sorted(BRITISH_TO_AMERICAN, key=lambda x: -len(x[0]))
]


def to_american(term: str) -> str:
    result = term
    for pattern, amer in _BRITISH_PATTERNS:
        result = pattern.sub(amer, result)
    return result


def display_term(american: str, _variants: list[str]) -> str:
    """American spelling, all lowercase."""
    return american.casefold()


def clean_term(text) -> str | None:
    if text is None:
        return None
    s = str(text).strip()
    if not s:
        return None
    s = re.sub(r"\binjection\s+site\b", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip(" ,;/-")
    return s or None


def collect_terms(wb) -> list[str]:
    raw: list[str] = []
    for sheet_name in wb.sheetnames:
        if sheet_name == "Glossary":
            continue
        ws = wb[sheet_name]
        for row in range(5, ws.max_row + 1):
            for col in (2, 4):
                cleaned = clean_term(ws.cell(row, col).value)
                if cleaned:
                    raw.append(cleaned)

    buckets: dict[str, list[str]] = {}
    for term in raw:
        key = to_american(term).casefold()
        buckets.setdefault(key, []).append(term)

    unique = [display_term(to_american(variants[0]), variants) for variants in buckets.values()]
    unique = [t for t in unique if t.casefold() not in EXCLUDED_TERMS]
    unique.sort(key=lambda x: x.casefold())
    return unique


def write_glossary_sheet(wb, terms: list[str]) -> None:
    if "Glossary" in wb.sheetnames:
        del wb["Glossary"]
    ws = wb.create_sheet("Glossary")
    ws["A1"] = "Glossary term"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = (
        "Columns B+D combined; injection site removed; "
        "case-insensitive and British/American spelling duplicates removed (American kept; all lowercase)"
    )
    for i, term in enumerate(terms, start=2):
        ws.cell(i, 1, term)


def main() -> None:
    if not DEFAULT_XLSX.exists():
        raise SystemExit(f"Workbook not found: {DEFAULT_XLSX}")

    wb = load_workbook(DEFAULT_XLSX)
    terms = collect_terms(wb)
    write_glossary_sheet(wb, terms)
    wb.save(DEFAULT_XLSX)

    payload = {
        "title": "MedDRA glossary",
        "source": (
            "MedDRA search term list — PTs and LLTs (columns B and D combined; "
            '"injection site" removed; case-insensitive and British/American duplicates removed; '
            "American spelling kept; all lowercase)"
        ),
        "terms": terms,
    }
    OUT_JS.write_text(
        "window.MEDDRA_GLOSSARY = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(terms)} terms to {OUT_JS}")
    print(f"Updated Glossary sheet in {DEFAULT_XLSX}")


if __name__ == "__main__":
    main()
