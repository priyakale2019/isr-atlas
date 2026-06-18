"""
Merge an Excel workbook into data/lesions.js for the injection-site morphology site.

Requirements:
  pip install openpyxl

Expected workbook layout (one sheet per category, or a single sheet — adjust SHEETS below):

Sheet "primary" (and optionally "secondary"):
  Columns:
    id              — stable slug, e.g. "macule"
    term            — display name
    definition      — clinical definition
    injectionContext — short injection-site note
    image_url       — primary example image URL or path under /assets/...
    image_caption
    image_credit
    image_license
    gallery_urls    — optional, pipe-separated list of extra image URLs
    gallery_captions — optional, pipe-separated parallel to gallery_urls

Relative paths in image_url are emitted as strings; the static site resolves them from the site root.

After editing, run:
  python scripts/import_excel_to_lesions.py path/to/workbook.xlsx

This overwrites ../data/lesions.js while preserving the narrative sections from the bundled template
if your workbook only contains image rows. For a full replacement, extend the script to read all fields.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError as exc:  # pragma: no cover - runtime helper
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DATA = ROOT / "data" / "lesions.js"


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def split_pipe(cell: object | None) -> list[str]:
    if cell is None or str(cell).strip() == "":
        return []
    return [part.strip() for part in str(cell).split("|") if part.strip()]


def apply_updates(base_list: list[dict], updates: list[dict]) -> list[dict]:
    """Preserve order and entries from base_list; overlay rows that share an id."""
    by_id = {entry["id"]: i for i, entry in enumerate(base_list)}
    out = [dict(entry) for entry in base_list]
    for upd in updates:
        if upd["id"] in by_id:
            out[by_id[upd["id"]]] = upd
        else:
            out.append(upd)
    return out


def merge_sheet_rows(ws, kind: str, base: dict) -> list[dict]:
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))[0:]]
    idx = {h: i for i, h in enumerate(headers) if h}
    rows_out: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        def col(name: str) -> object | None:
            if name not in idx:
                return None
            i = idx[name]
            return row[i] if i < len(row) else None

        term = col("term")
        if not term:
            continue
        term_s = str(term).strip()
        ident = col("id")
        ident_s = str(ident).strip() if ident else slugify(term_s)

        image = {
            "url": str(col("image_url") or "").strip(),
            "caption": str(col("image_caption") or term_s).strip(),
            "credit": str(col("image_credit") or "Unknown").strip(),
            "license": str(col("image_license") or "License not stated").strip(),
        }

        gallery_urls = split_pipe(col("gallery_urls"))
        gallery_caps = split_pipe(col("gallery_captions"))
        gallery: list[dict] = []
        for i, url in enumerate(gallery_urls):
            cap = gallery_caps[i] if i < len(gallery_caps) else f"{term_s} — additional figure"
            gallery.append(
                {
                    "url": url,
                    "caption": cap,
                    "credit": str(col("image_credit") or "").strip() or "Unknown",
                    "license": str(col("image_license") or "").strip() or "License not stated",
                }
            )

        entry = {
            "id": ident_s,
            "term": term_s,
            "definition": str(col("definition") or "").strip(),
            "injectionContext": str(col("injectionContext") or col("injection_context") or "").strip(),
            "images": [image],
            "gallery": gallery,
        }

        # Preserve prior definition/injection text if Excel cells left blank.
        existing = next((x for x in base[kind] if x["id"] == ident_s), None)
        if existing:
            if not entry["definition"]:
                entry["definition"] = existing["definition"]
            if not entry["injectionContext"]:
                entry["injectionContext"] = existing["injectionContext"]
            if not image["url"]:
                entry["images"] = existing["images"]
            if not gallery and existing.get("gallery"):
                entry["gallery"] = existing["gallery"]

        rows_out.append(entry)

    return rows_out


def load_base_data() -> dict:
    text = DEFAULT_DATA.read_text(encoding="utf-8")
    if not text.startswith("window.LESION_DATA"):
        raise SystemExit("data/lesions.js missing window.LESION_DATA")
    payload = text.split("=", 1)[1].strip().rstrip(";")
    return json.loads(payload)


def main(argv: list[str]) -> None:
    if len(argv) < 2:
        raise SystemExit("Usage: python import_excel_to_lesions.py workbook.xlsx")

    workbook_path = Path(argv[1]).expanduser().resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True)
    base = load_base_data()

    if "primary" in wb.sheetnames:
        merged = merge_sheet_rows(wb["primary"], "primaryLesions", base)
        if merged:
            base["primaryLesions"] = apply_updates(base["primaryLesions"], merged)

    if "secondary" in wb.sheetnames:
        merged = merge_sheet_rows(wb["secondary"], "secondaryLesions", base)
        if merged:
            base["secondaryLesions"] = apply_updates(base["secondaryLesions"], merged)

    out = "window.LESION_DATA = " + json.dumps(base, ensure_ascii=False, indent=2) + ";\n"
    DEFAULT_DATA.write_text(out, encoding="utf-8")
    print(f"Wrote {DEFAULT_DATA}")


if __name__ == "__main__":
    main(sys.argv)
