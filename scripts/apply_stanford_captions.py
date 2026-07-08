"""
Apply Stanford Dermatology Physician captions (and matching photos) from the
Website pictures workbook to primary morphology entries in data/lesions.js.

Usage:
  python scripts/apply_stanford_captions.py "/path/to/Copy of Gilead Atlas - Website pictures (1).xlsx"
"""

from __future__ import annotations

import json
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
LESIONS_PATH = ROOT / "data" / "lesions.js"
ASSET_MEDIA = ROOT / "assets" / "atlas" / "media"
NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"

PRIMARY_FEATURES = {
    "macule",
    "patch",
    "papule",
    "plaque",
    "nodule",
    "ulcer",
    "abscess",
    "phlebitis",
    "ecchymosis",
    "pustule",
    "vesicle",
    "bulla",
}


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "feature"


def parse_drawing_anchors(z: zipfile.ZipFile, drawing_path: str, rels_path: str) -> dict[tuple[int, int], str]:
    root = ET.fromstring(z.read(drawing_path))
    rels = ET.fromstring(z.read(rels_path))
    rid_to_media = {
        rel.attrib["Id"]: rel.attrib["Target"].split("/")[-1]
        for rel in rels
        if "image" in rel.attrib.get("Type", "")
    }
    mapping: dict[tuple[int, int], str] = {}
    for anchor in root:
        if not anchor.tag.endswith("CellAnchor"):
            continue
        from_cell = anchor.find(f"{{{NS_XDR}}}from")
        if from_cell is None:
            continue
        row = int(from_cell.find(f"{{{NS_XDR}}}row").text) + 1
        col = int(from_cell.find(f"{{{NS_XDR}}}col").text) + 1
        embed = None
        for blip in anchor.iter():
            if blip.tag.endswith("}blip"):
                for k, v in blip.attrib.items():
                    if k.endswith("embed"):
                        embed = v
        if embed and embed in rid_to_media:
            mapping[(row, col)] = rid_to_media[embed]
    return mapping


def extract_media_zip(xlsx: Path, dest: Path) -> None:
    dest.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(xlsx) as z:
        for name in z.namelist():
            if not name.startswith("xl/media/"):
                continue
            out = dest / Path(name).name
            out.write_bytes(z.read(name))


def picture_number(label: str) -> int:
    m = re.search(r"picture\s*(\d+)", label.lower())
    if not m:
        raise ValueError(f"Could not parse picture number from {label!r}")
    return int(m.group(1))


def load_rows(xlsx: Path) -> dict[str, list[dict]]:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb.active

    with zipfile.ZipFile(xlsx) as z:
        drawing = "xl/drawings/drawing1.xml"
        rels = "xl/drawings/_rels/drawing1.xml.rels"
        anchors = parse_drawing_anchors(z, drawing, rels)

    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        feature = ws.cell(row=row, column=1).value
        if not feature:
            continue
        feature_key = str(feature).strip().lower()
        if feature_key not in PRIMARY_FEATURES:
            continue

        label = ws.cell(row=row, column=3).value
        if not label:
            continue
        media = anchors.get((row, 4))
        if not media:
            continue

        stanford = ws.cell(row=row, column=6).value
        caption = str(stanford).strip() if stanford else ""
        if not caption:
            published = ws.cell(row=row, column=5).value
            caption = str(published).strip() if published else f"{feature} — atlas figure"

        grouped[feature_key].append(
            {
                "picture": picture_number(str(label)),
                "url": f"assets/atlas/media/{media}",
                "caption": caption,
                "credit": "",
                "license": "",
            }
        )

    for shots in grouped.values():
        shots.sort(key=lambda s: s["picture"])
        for shot in shots:
            shot.pop("picture", None)
    return grouped


def main() -> None:
    xlsx = Path(sys.argv[1] if len(sys.argv) > 1 else "")
    if not xlsx.exists():
        raise SystemExit(f"Workbook not found: {xlsx}")

    extract_media_zip(xlsx, ASSET_MEDIA)
    grouped = load_rows(xlsx)

    text = LESIONS_PATH.read_text(encoding="utf-8")
    if not text.startswith("window.LESION_DATA"):
        raise SystemExit("Unexpected lesions.js format")
    data = json.loads(text.split("=", 1)[1].strip().rstrip(";"))

    updated = []
    for entry in data["primaryLesions"]:
        ident = entry["id"]
        shots = grouped.get(ident, [])
        if not shots:
            updated.append(entry)
            continue
        entry = dict(entry)
        entry["images"] = [shots[0]]
        entry["gallery"] = shots[1:]
        updated.append(entry)

    data["primaryLesions"] = updated
    out = "window.LESION_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n"
    LESIONS_PATH.write_text(out, encoding="utf-8")
    print(f"Updated {LESIONS_PATH}")
    for ident in PRIMARY_FEATURES:
        count = len(grouped.get(ident, []))
        print(f"  {ident}: {count} photo(s)")


if __name__ == "__main__":
    main()
