"""
Extract embedded images from the Gilead ISR Atlas Excel workbook into assets/atlas/media/
and regenerate data/lesions.js for the static site.

Default source paths (override with env GILEAD_ATLAS_XLSX):
  Dropbox ISRWebsite / Proposed Gilead Atlas - Draft #3.xlsx

Optional: copy the companion PDF to assets/atlas/ when GILEAD_ATLAS_PDF is set.
"""

from __future__ import annotations

import json
import os
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ASSET_MEDIA = ROOT / "assets" / "atlas" / "media"
ASSET_ATLAS = ROOT / "assets" / "atlas"

DEFAULT_XLSX = Path(
    r"C:\Users\ksarin\Dropbox\KavitaScripts\ISRWebsite\Proposed Gilead Atlas - Draft #3.xlsx"
)
DEFAULT_PDF = Path(
    r"C:\Users\ksarin\Dropbox\KavitaScripts\ISRWebsite\Proposed Gilead Atlas - Draft #3.pdf"
)

NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "feature"


def parse_drawing_anchors(z: zipfile.ZipFile, drawing_path: str, rels_path: str) -> list[tuple[int, int, str]]:
    root = ET.fromstring(z.read(drawing_path))
    rels = {rel.attrib["Id"]: rel.attrib["Target"] for rel in ET.fromstring(z.read(rels_path))}
    rid_to_media: dict[str, str] = {}
    for rid, tgt in rels.items():
        if "media/" in tgt:
            rid_to_media[rid] = tgt.split("/")[-1]

    ordered: list[tuple[int, int, str]] = []
    for anchor in root:
        if not anchor.tag.endswith("CellAnchor"):
            continue
        from_cell = anchor.find(f"{{{NS_XDR}}}from")
        if from_cell is None:
            continue
        row_el = from_cell.find(f"{{{NS_XDR}}}row")
        col_el = from_cell.find(f"{{{NS_XDR}}}col")
        if row_el is None or col_el is None:
            continue
        excel_row = int(row_el.text) + 1
        excel_col = int(col_el.text) + 1
        embed = None
        for blip in anchor.iter():
            if blip.tag.endswith("}blip"):
                for k, v in blip.attrib.items():
                    if k.endswith("embed"):
                        embed = v
                        break
        if embed and embed in rid_to_media:
            ordered.append((excel_row, excel_col, rid_to_media[embed]))
    return ordered


def row_images_from_anchors(anchors: list[tuple[int, int, str]], row: int) -> list[str]:
    """Preserve document order for images on a row; drop consecutive duplicate files."""
    media: list[str] = []
    for r, _c, m in anchors:
        if r != row:
            continue
        if media and media[-1] == m:
            continue
        media.append(m)
    return media


def caption_for_picture_slot(ws, row: int, picture_col: int) -> str:
    """Picture N is in column picture_col; notes in picture_col + 1."""
    notes = ws.cell(row=row, column=picture_col + 1).value
    if notes and str(notes).strip():
        return str(notes).strip().replace("\r\n", "\n")
    feat = ws.cell(row=row, column=1).value
    return f"{feat} — atlas figure"


def credit_for_picture_slot(ws, row: int, picture_col: int, reference_col_offset: int) -> str:
    """APA / source line when reference_col_offset is set (website sheet)."""
    if reference_col_offset <= 0:
        return ""
    ref = ws.cell(row=row, column=picture_col + reference_col_offset).value
    return str(ref).strip() if ref else ""


def extract_media_zip(xlsx: Path, dest: Path) -> None:
    dest.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(xlsx) as z:
        for name in z.namelist():
            if not name.startswith("xl/media/"):
                continue
            base = Path(name).name
            out = dest / base
            out.write_bytes(z.read(name))


def build_lesion_entries(
    ws,
    anchors: list[tuple[int, int, str]],
    header_row: int,
    credit: str,
    license_line: str,
    reference_col_offset: int = 0,
) -> list[dict]:
    entries: list[dict] = []
    for row in range(header_row + 1, ws.max_row + 1):
        feature = ws.cell(row=row, column=1).value
        if not feature or str(feature).strip() in ("Feature", "primary morphology", "secondary morphology"):
            continue
        definition = ws.cell(row=row, column=2).value
        definition_s = str(definition).strip() if definition else ""
        if not row_images_from_anchors(anchors, row) and not definition_s:
            continue
        term = str(feature).strip()
        ident = slugify(term)
        images: list[dict] = []
        gallery: list[dict] = []
        # Map each embedded image to the column it anchors on for caption lookup
        raw_cols = [(c, m) for r, c, m in anchors if r == row]
        cols_for_row = [pair for _, pair in sorted(enumerate(raw_cols), key=lambda im: (im[1][0], im[0]))]
        for idx, (col, mf) in enumerate(cols_for_row):
            cap = caption_for_picture_slot(ws, row, col)
            shot_credit = credit_for_picture_slot(ws, row, col, reference_col_offset) or credit
            shot = {
                "url": f"assets/atlas/media/{mf}",
                "caption": cap,
                "credit": shot_credit,
                "license": license_line,
            }
            if idx == 0:
                images.append(shot)
            else:
                gallery.append(shot)
        inj = (
            "Use when documenting local findings at or near an injection site, consistent with your "
            "institutional adverse-event and photography policies."
        )
        entries.append(
            {
                "id": ident,
                "term": term[0].upper() + term[1:] if term else ident,
                "definition": definition_s or f"Atlas entry: {term}.",
                "injectionContext": inj,
                "images": images or [],
                "gallery": gallery,
            }
        )
    return entries


def main() -> None:
    xlsx = Path(os.environ.get("GILEAD_ATLAS_XLSX", str(DEFAULT_XLSX)))
    pdf = Path(os.environ.get("GILEAD_ATLAS_PDF", str(DEFAULT_PDF)))
    if not xlsx.exists():
        raise SystemExit(f"Workbook not found: {xlsx}")

    ASSET_MEDIA.mkdir(parents=True, exist_ok=True)
    extract_media_zip(xlsx, ASSET_MEDIA)

    if pdf.exists():
        ASSET_ATLAS.mkdir(parents=True, exist_ok=True)
        shutil.copy2(pdf, ASSET_ATLAS / "Proposed-Gilead-Atlas-Draft-3.pdf")

    use_website_primary = os.environ.get("GILEAD_PRIMARY_SOURCE", "website").lower() != "atlas"
    use_website_secondary = os.environ.get("GILEAD_SECONDARY_SOURCE", "website").lower() != "atlas"

    z = zipfile.ZipFile(xlsx)
    if use_website_primary:
        primary_anchors = parse_drawing_anchors(
            z, "xl/drawings/drawing2.xml", "xl/drawings/_rels/drawing2.xml.rels"
        )
    else:
        primary_anchors = parse_drawing_anchors(
            z, "xl/drawings/drawing1.xml", "xl/drawings/_rels/drawing1.xml.rels"
        )
    if use_website_secondary:
        secondary_anchors = parse_drawing_anchors(
            z, "xl/drawings/drawing4.xml", "xl/drawings/_rels/drawing4.xml.rels"
        )
    else:
        secondary_anchors = parse_drawing_anchors(
            z, "xl/drawings/drawing3.xml", "xl/drawings/_rels/drawing3.xml.rels"
        )
    z.close()

    wb = load_workbook(xlsx, data_only=True)
    credit = ""
    license_line = ""

    if use_website_primary:
        primary_ws = wb["Primary morphology - Website"]
        primary_ref_offset = 0  # captions only; APA reference columns omitted on site
    else:
        primary_ws = wb["primary morphology"]
        primary_ref_offset = 0

    if use_website_secondary:
        secondary_ws = wb["Secondary morphology - Website"]
    else:
        secondary_ws = wb["secondary morphology"]

    primary_entries = build_lesion_entries(
        primary_ws,
        primary_anchors,
        header_row=3,
        credit=credit,
        license_line=license_line,
        reference_col_offset=primary_ref_offset,
    )
    secondary_entries = build_lesion_entries(
        secondary_ws,
        secondary_anchors,
        header_row=2,
        credit=credit,
        license_line=license_line,
        reference_col_offset=0,
    )

    # Drop rows with no images and no definition (stray blanks)
    primary_entries = [e for e in primary_entries if e["images"] or e["definition"]]
    secondary_entries = [e for e in secondary_entries if e["images"] or e["definition"]]

    lesions_path = ROOT / "data" / "lesions.js"
    existing_text = lesions_path.read_text(encoding="utf-8")
    if not existing_text.startswith("window.LESION_DATA"):
        raise SystemExit("Unexpected lesions.js format")
    base = json.loads(existing_text.split("=", 1)[1].strip().rstrip(";"))

    base["meta"]["subtitle"] = (
        "Structured language for injection-site findings — clinical photographs with line-drawing references."
    )
    base["meta"]["atlasPdf"] = "assets/atlas/Proposed-Gilead-Atlas-Draft-3.pdf"
    base["meta"]["disclaimer"] = (
        "This resource is for clinical education and standardized description of cutaneous findings. "
        "Confirm image use, consent, and copyright with your organization before wider distribution. "
        "It does not establish diagnosis or management."
    )
    base["primaryLesions"] = primary_entries
    base["secondaryLesions"] = secondary_entries

    out = "window.LESION_DATA = " + json.dumps(base, ensure_ascii=False, indent=2) + ";\n"
    lesions_path.write_text(out, encoding="utf-8")
    print(f"Wrote {lesions_path}")
    print(f"Extracted media to {ASSET_MEDIA} ({len(list(ASSET_MEDIA.glob('*')))} files)")
    if (ASSET_ATLAS / "Proposed-Gilead-Atlas-Draft-3.pdf").exists():
        print(f"Copied PDF to {ASSET_ATLAS / 'Proposed-Gilead-Atlas-Draft-3.pdf'}")


if __name__ == "__main__":
    main()
